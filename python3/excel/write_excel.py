#!/usr/bin/python3

import openpyxl

# 1. 读取excel文档
wb = openpyxl.load_workbook('demo.xlsx')

sheet = wb['Sheet1']
# sheet的属性
# print(sheet.max_column)
# print(sheet.max_row)
# print(sheet.title)
sheet.title = 'example'
# print(sheet.title)

wb.save(filename="example.xlsx")
