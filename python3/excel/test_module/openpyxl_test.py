import openpyxl


def test_openpyxl_read():
    f = openpyxl.load_workbook('test_cases\\read_xlsx.xlsx', read_only=True)


c = [65536] * 1200
f = openpyxl.Workbook(write_only=True)


def test_openpyxl_write():
    for i in range(1):
        sheet = f.create_sheet(title=str(i))
        for row in range(2000):
            sheet.append(c)
    f.save('test_cases\\write_xlsx.xlsx')


if __name__ == '__main__':
    test_openpyxl_write()